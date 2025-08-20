from openpyxl import load_workbook
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.widgets import Slider, RadioButtons
from scipy import stats, spatial
from scipy.ndimage import gaussian_filter
import seaborn as sns
from typing import Dict, List, Tuple, Optional, Union

def load_excel_data(file_path: str) -> Optional[pd.DataFrame]:
    """
    Load data from an Excel file with enhanced error handling and data validation.
    
    Parameters:
    -----------
    file_path : str
        Path to the Excel file
        
    Returns:
    --------
    pd.DataFrame or None
        Loaded data or None if failed
    """
    try:
        # Try to load with pandas first for better header handling
        df = pd.read_excel(file_path, engine='openpyxl')
        
        # Data cleaning and validation
        df = clean_geological_data(df)
        
        return df
    except Exception as e:
        print(f"Error loading Excel file with pandas: {e}")
        try:
            # Fallback to openpyxl
            workbook = load_workbook(filename=file_path)
            sheet = workbook.active
            
            # Get header row
            headers = [cell.value for cell in sheet[1]]
            
            # Get data rows
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data.append(row)
                
            df = pd.DataFrame(data, columns=headers)
            df = clean_geological_data(df)
            
            return df
        except Exception as e2:
            print(f"Error loading Excel file with openpyxl: {e2}")
            return None

def clean_geological_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean and standardize geological data.
    
    Parameters:
    -----------
    df : pd.DataFrame
        Raw geological data
        
    Returns:
    --------
    pd.DataFrame
        Cleaned data
    """
    if df is None or df.empty:
        return df
        
    # Remove completely empty rows and columns
    df = df.dropna(how='all').dropna(axis=1, how='all')
    
    # Standardize column names (remove spaces, convert to consistent case)
    df.columns = df.columns.str.strip().str.replace(' ', '_')
    
    # Convert numeric columns to appropriate types
    numeric_cols = ['PK_medio', 'RMR', 'Buzamiento', 'Direccion_Buzamiento', 'X', 'Y']
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # Handle missing values in key columns
    if 'RMR' in df.columns:
        df['RMR'] = df['RMR'].fillna(0)
        df['RMR'] = df['RMR'].clip(0, 100)  # Ensure RMR is within valid range
    
    if 'Buzamiento' in df.columns:
        df['Buzamiento'] = df['Buzamiento'].fillna(0)
        df['Buzamiento'] = df['Buzamiento'].clip(0, 90)  # Dip angle 0-90 degrees
    
    if 'Direccion_Buzamiento' in df.columns:
        df['Direccion_Buzamiento'] = df['Direccion_Buzamiento'].fillna(0)
        df['Direccion_Buzamiento'] = df['Direccion_Buzamiento'] % 360  # Normalize to 0-360
    
    return df

def analyze_data(data):
    """Perform basic data analysis on the loaded data."""
    if data is None:
        return None
    
    # Example analysis: calculate mean and standard deviation
    analysis_results = {
        'mean': data.mean().to_dict() if hasattr(data.mean(), 'to_dict') else str(data.mean()),
        'std_dev': data.std().to_dict() if hasattr(data.std(), 'to_dict') else str(data.std()),
        'count': len(data),
        'columns': data.columns.tolist()
    }
    return analysis_results

def filter_data(data, column_name, threshold):
    """Filter data based on a threshold for a specific column."""
    if column_name in data.columns:
        filtered_data = data[data[column_name] > threshold]
        return filtered_data
    else:
        print(f"Column {column_name} does not exist in the data.")
        return None

def filter_geological_data(data, filters):
    """Filter geological data based on multiple criteria."""
    if not data:
        return []
    
    df = pd.DataFrame(data)
    filtered_df = df.copy()
    
    for filter_key, filter_value in filters.items():
        if filter_key in df.columns and filter_value is not None:
            if isinstance(filter_value, dict):
                # Range filter
                if 'min' in filter_value and filter_value['min'] is not None:
                    filtered_df = filtered_df[filtered_df[filter_key] >= filter_value['min']]
                if 'max' in filter_value and filter_value['max'] is not None:
                    filtered_df = filtered_df[filtered_df[filter_key] <= filter_value['max']]
            else:
                # Exact match filter
                filtered_df = filtered_df[filtered_df[filter_key] == filter_value]
    
    return filtered_df.to_dict('records')

def analyze_rmr_data(data, filters):
    """Analyze RMR (Rock Mass Rating) data with geological criteria."""
    if not data:
        return []
    
    df = pd.DataFrame(data)
    
    # Apply filters
    filtered_data = filter_geological_data(data, filters)
    filtered_df = pd.DataFrame(filtered_data)
    
    if filtered_df.empty:
        return []
    
    # Calculate RMR coordinates and colors based on rating
    rmr_data = []
    for _, row in filtered_df.iterrows():
        rmr_value = row.get('RMR', 0) if 'RMR' in row else 0
        
        # Determine color based on RMR value
        if rmr_value >= 80:
            color = '#00ff00'  # Green - Very good rock
            class_name = 'Muy Buena'
        elif rmr_value >= 60:
            color = '#80ff00'  # Light green - Good rock
            class_name = 'Buena'
        elif rmr_value >= 40:
            color = '#ffff00'  # Yellow - Fair rock
            class_name = 'Regular'
        elif rmr_value >= 20:
            color = '#ff8000'  # Orange - Poor rock
            class_name = 'Mala'
        else:
            color = '#ff0000'  # Red - Very poor rock
            class_name = 'Muy Mala'
        
        rmr_record = row.to_dict()
        rmr_record.update({
            'color': color,
            'class': class_name,
            'x': row.get('X', 0),
            'y': row.get('Y', 0),
            'rmr_value': rmr_value
        })
        rmr_data.append(rmr_record)
    
    return rmr_data

def analyze_fracture_data(data, filters):
    """Analyze fracture data for structural geology visualization."""
    if not data:
        return []
    
    df = pd.DataFrame(data)
    
    # Apply filters
    filtered_data = filter_geological_data(data, filters)
    filtered_df = pd.DataFrame(filtered_data)
    
    if filtered_df.empty:
        return []
    
    # Group fractures by family and assign colors
    fracture_families = {}
    family_colors = ['#ff0000', '#00ff00', '#0000ff', '#ffff00', '#ff00ff', '#00ffff']
    
    fracture_data = []
    for _, row in filtered_df.iterrows():
        family = row.get('Familia', 'Unknown')
        if family not in fracture_families:
            color_index = len(fracture_families) % len(family_colors)
            fracture_families[family] = family_colors[color_index]
        
        fracture_record = row.to_dict()
        fracture_record.update({
            'color': fracture_families[family],
            'family': family,
            'dip': row.get('Buzamiento', 0),
            'dip_direction': row.get('Direccion_Buzamiento', 0),
            'x': row.get('X', 0),
            'y': row.get('Y', 0)
        })
        fracture_data.append(fracture_record)
    
    return fracture_data

def calculate_excavation_coordinates(data, image_dimensions, scale_factor=1.0):
    """Calculate coordinates for plotting on image based on PK and frente data."""
    coordinates = []
    
    for record in data:
        pk = record.get('PK_medio', 0)
        frente = record.get('Frente', 0)
        
        # Convert geological coordinates to image pixels
        # This is a simplified conversion - in reality, you'd need proper coordinate transformation
        x = int((pk * scale_factor) % image_dimensions['width'])
        y = int((frente * scale_factor) % image_dimensions['height'])
        
        record['image_x'] = x
        record['image_y'] = y
        coordinates.append(record)
    
    return coordinates

def generate_visualization(data: Union[pd.DataFrame, List[Dict]]) -> Dict:
    """
    Generate comprehensive visualizations based on the analyzed data.
    
    Parameters:
    -----------
    data : pd.DataFrame or list of dicts
        Geological data to visualize
        
    Returns:
    --------
    dict
        Visualization data and statistics
    """
    visualization_data = {
        'charts': [],
        'statistics': {},
        'recommendations': [],
        'interactive_plots': {}
    }
    
    if data is None:
        return visualization_data
    
    if isinstance(data, list):
        df = pd.DataFrame(data)
    else:
        df = data.copy()
    
    if df.empty:
        return visualization_data
    
    # Basic statistics
    visualization_data['statistics'] = {
        'total_records': len(df),
        'columns': df.columns.tolist(),
        'summary': df.describe().to_dict() if len(df) > 0 else {}
    }
    
    # Generate various visualizations
    try:
        # RMR distribution if available
        if 'RMR' in df.columns:
            rmr_stats = calculate_rmr_statistics(df)
            visualization_data['statistics']['rmr'] = rmr_stats
            visualization_data['charts'].append('rmr_distribution')
        
        # Fracture analysis if available
        if 'Familia' in df.columns and 'Buzamiento' in df.columns:
            fracture_stats = calculate_fracture_statistics(df)
            visualization_data['statistics']['fractures'] = fracture_stats
            visualization_data['charts'].append('fracture_rosette')
        
        # Spatial analysis if coordinates available
        if all(col in df.columns for col in ['X', 'Y']):
            spatial_stats = calculate_spatial_statistics(df)
            visualization_data['statistics']['spatial'] = spatial_stats
            visualization_data['charts'].append('spatial_distribution')
        
        # Generate recommendations
        visualization_data['recommendations'] = generate_geological_recommendations(df)
        
    except Exception as e:
        print(f"Error generating visualizations: {e}")
    
    return visualization_data

def calculate_rmr_statistics(df: pd.DataFrame) -> Dict:
    """Calculate RMR-specific statistics."""
    rmr_data = df['RMR'].dropna()
    
    if rmr_data.empty:
        return {}
    
    stats_dict = {
        'mean': float(rmr_data.mean()),
        'std': float(rmr_data.std()),
        'min': float(rmr_data.min()),
        'max': float(rmr_data.max()),
        'percentiles': {
            'p25': float(rmr_data.quantile(0.25)),
            'p50': float(rmr_data.median()),
            'p75': float(rmr_data.quantile(0.75))
        }
    }
    
    # RMR classification distribution
    classifications = []
    for rmr in rmr_data:
        if rmr >= 80:
            classifications.append('Very Good')
        elif rmr >= 60:
            classifications.append('Good')
        elif rmr >= 40:
            classifications.append('Fair')
        elif rmr >= 20:
            classifications.append('Poor')
        else:
            classifications.append('Very Poor')
    
    classification_counts = pd.Series(classifications).value_counts()
    stats_dict['classification_distribution'] = classification_counts.to_dict()
    
    return stats_dict

def calculate_fracture_statistics(df: pd.DataFrame) -> Dict:
    """Calculate fracture-specific statistics."""
    fracture_cols = ['Familia', 'Buzamiento', 'Direccion_Buzamiento']
    available_cols = [col for col in fracture_cols if col in df.columns]
    
    if not available_cols:
        return {}
    
    stats_dict = {}
    
    # Family distribution
    if 'Familia' in df.columns:
        family_counts = df['Familia'].value_counts()
        stats_dict['family_distribution'] = family_counts.to_dict()
    
    # Dip statistics
    if 'Buzamiento' in df.columns:
        dip_data = df['Buzamiento'].dropna()
        if not dip_data.empty:
            stats_dict['dip'] = {
                'mean': float(dip_data.mean()),
                'std': float(dip_data.std()),
                'dominant_range': get_dominant_dip_range(dip_data)
            }
    
    # Dip direction statistics (orientation analysis)
    if 'Direccion_Buzamiento' in df.columns:
        direction_data = df['Direccion_Buzamiento'].dropna()
        if not direction_data.empty:
            stats_dict['dip_direction'] = {
                'mean_vector': calculate_circular_mean(direction_data),
                'dispersion': calculate_circular_dispersion(direction_data),
                'dominant_sets': identify_joint_sets(direction_data)
            }
    
    return stats_dict

def calculate_spatial_statistics(df: pd.DataFrame) -> Dict:
    """Calculate spatial distribution statistics."""
    coords = df[['X', 'Y']].dropna()
    
    if coords.empty:
        return {}
    
    stats_dict = {
        'extent': {
            'x_min': float(coords['X'].min()),
            'x_max': float(coords['X'].max()),
            'y_min': float(coords['Y'].min()),
            'y_max': float(coords['Y'].max())
        },
        'centroid': {
            'x': float(coords['X'].mean()),
            'y': float(coords['Y'].mean())
        }
    }
    
    # Calculate spatial clustering if enough points
    if len(coords) > 10:
        try:
            # Calculate nearest neighbor distances
            tree = spatial.cKDTree(coords.values)
            distances, _ = tree.query(coords.values, k=2)
            nn_distances = distances[:, 1]  # First neighbor (exclude self)
            
            stats_dict['clustering'] = {
                'mean_nn_distance': float(np.mean(nn_distances)),
                'std_nn_distance': float(np.std(nn_distances)),
                'clustering_index': calculate_clustering_index(nn_distances)
            }
        except Exception as e:
            print(f"Error calculating spatial clustering: {e}")
    
    return stats_dict

def get_dominant_dip_range(dip_data: pd.Series) -> str:
    """Identify dominant dip range."""
    if dip_data.empty:
        return "No data"
    
    mean_dip = dip_data.mean()
    
    if mean_dip < 30:
        return "Shallow (0-30°)"
    elif mean_dip < 60:
        return "Moderate (30-60°)"
    else:
        return "Steep (60-90°)"

def calculate_circular_mean(angles: pd.Series) -> float:
    """Calculate circular mean for angular data."""
    angles_rad = np.radians(angles)
    sin_sum = np.sum(np.sin(angles_rad))
    cos_sum = np.sum(np.cos(angles_rad))
    mean_angle = np.arctan2(sin_sum, cos_sum)
    return float(np.degrees(mean_angle) % 360)

def calculate_circular_dispersion(angles: pd.Series) -> float:
    """Calculate circular dispersion for angular data."""
    angles_rad = np.radians(angles)
    sin_sum = np.sum(np.sin(angles_rad))
    cos_sum = np.sum(np.cos(angles_rad))
    r = np.sqrt(sin_sum**2 + cos_sum**2) / len(angles)
    return float(1 - r)

def identify_joint_sets(directions: pd.Series) -> List[Dict]:
    """Identify major joint sets from dip direction data."""
    if directions.empty:
        return []
    
    # Simple binning approach - could be enhanced with clustering
    bins = np.arange(0, 361, 30)  # 30-degree bins
    hist, bin_edges = np.histogram(directions, bins=bins)
    
    joint_sets = []
    for i, count in enumerate(hist):
        if count > len(directions) * 0.1:  # At least 10% of data
            bin_center = (bin_edges[i] + bin_edges[i+1]) / 2
            joint_sets.append({
                'direction': float(bin_center),
                'count': int(count),
                'percentage': float(count / len(directions) * 100)
            })
    
    return sorted(joint_sets, key=lambda x: x['count'], reverse=True)

def calculate_clustering_index(distances: np.ndarray) -> float:
    """Calculate clustering index based on nearest neighbor distances."""
    observed_mean = np.mean(distances)
    # Expected mean for random distribution would be 1/(2*sqrt(density))
    # This is a simplified index
    return float(observed_mean / np.std(distances))

def generate_geological_recommendations(df: pd.DataFrame) -> List[str]:
    """Generate geological engineering recommendations based on data analysis."""
    recommendations = []
    
    # RMR-based recommendations
    if 'RMR' in df.columns:
        rmr_data = df['RMR'].dropna()
        if not rmr_data.empty:
            mean_rmr = rmr_data.mean()
            
            if mean_rmr >= 80:
                recommendations.append("Excellent rock conditions. Minimal support required.")
            elif mean_rmr >= 60:
                recommendations.append("Good rock conditions. Light support recommended.")
            elif mean_rmr >= 40:
                recommendations.append("Fair rock conditions. Systematic support required.")
            elif mean_rmr >= 20:
                recommendations.append("Poor rock conditions. Heavy support and close monitoring needed.")
            else:
                recommendations.append("Very poor rock conditions. Immediate and extensive support required.")
            
            # Variability check
            if rmr_data.std() > 20:
                recommendations.append("High RMR variability detected. Consider variable support design.")
    
    # Fracture analysis recommendations
    if 'Familia' in df.columns:
        family_counts = df['Familia'].value_counts()
        if len(family_counts) >= 3:
            recommendations.append("Multiple fracture families present. Consider wedge stability analysis.")
    
    # Spatial distribution recommendations
    if all(col in df.columns for col in ['X', 'Y']):
        coords = df[['X', 'Y']].dropna()
        if not coords.empty and len(coords) > 5:
            x_range = coords['X'].max() - coords['X'].min()
            y_range = coords['Y'].max() - coords['Y'].min()
            
            if max(x_range, y_range) / min(x_range, y_range) > 3:
                recommendations.append("Elongated spatial distribution. Consider directional effects.")
    
    return recommendations

def create_interactive_rmr_plot(df: pd.DataFrame) -> plt.Figure:
    """
    Create interactive RMR plot with sliders for filtering.
    
    Parameters:
    -----------
    df : pd.DataFrame
        Geological data with RMR values
        
    Returns:
    --------
    matplotlib.figure.Figure
        Interactive RMR plot
    """
    if 'RMR' not in df.columns:
        return None
    
    fig, ax = plt.subplots(figsize=(12, 8))
    plt.subplots_adjust(bottom=0.25, left=0.15)
    
    # Initial plot
    rmr_data = df['RMR'].dropna()
    coords = df[['X', 'Y']].dropna() if all(col in df.columns for col in ['X', 'Y']) else None
    
    if coords is not None and len(coords) == len(rmr_data):
        scatter = ax.scatter(coords['X'], coords['Y'], c=rmr_data, cmap='RdYlGn', 
                           s=50, alpha=0.7, vmin=0, vmax=100)
        ax.set_xlabel('X Coordinate')
        ax.set_ylabel('Y Coordinate')
        ax.set_title('RMR Spatial Distribution')
        plt.colorbar(scatter, ax=ax, label='RMR Value')
    else:
        # Histogram if no spatial coordinates
        ax.hist(rmr_data, bins=20, alpha=0.7, color='skyblue', edgecolor='black')
        ax.set_xlabel('RMR Value')
        ax.set_ylabel('Frequency')
        ax.set_title('RMR Distribution')
    
    # Create sliders for filtering
    ax_min_rmr = plt.axes([0.15, 0.1, 0.3, 0.03])
    ax_max_rmr = plt.axes([0.55, 0.1, 0.3, 0.03])
    
    slider_min = Slider(ax_min_rmr, 'Min RMR', 0, 100, valinit=rmr_data.min())
    slider_max = Slider(ax_max_rmr, 'Max RMR', 0, 100, valinit=rmr_data.max())
    
    # Update function for sliders
    def update_rmr_plot(val):
        min_rmr = slider_min.val
        max_rmr = slider_max.val
        
        # Filter data
        mask = (rmr_data >= min_rmr) & (rmr_data <= max_rmr)
        filtered_rmr = rmr_data[mask]
        
        ax.clear()
        
        if coords is not None:
            filtered_coords = coords[mask]
            if not filtered_coords.empty:
                scatter = ax.scatter(filtered_coords['X'], filtered_coords['Y'], 
                                   c=filtered_rmr, cmap='RdYlGn', s=50, alpha=0.7, 
                                   vmin=0, vmax=100)
                ax.set_xlabel('X Coordinate')
                ax.set_ylabel('Y Coordinate')
        else:
            ax.hist(filtered_rmr, bins=20, alpha=0.7, color='skyblue', edgecolor='black')
            ax.set_xlabel('RMR Value')
            ax.set_ylabel('Frequency')
        
        ax.set_title(f'RMR Distribution (Range: {min_rmr:.0f}-{max_rmr:.0f})')
        fig.canvas.draw()
    
    slider_min.on_changed(update_rmr_plot)
    slider_max.on_changed(update_rmr_plot)
    
    return fig

def create_interactive_fracture_plot(df: pd.DataFrame) -> plt.Figure:
    """
    Create interactive fracture analysis plot with radio buttons for family selection.
    
    Parameters:
    -----------
    df : pd.DataFrame
        Geological data with fracture information
        
    Returns:
    --------
    matplotlib.figure.Figure
        Interactive fracture plot
    """
    if 'Familia' not in df.columns:
        return None
    
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))
    plt.subplots_adjust(left=0.25)
    
    families = df['Familia'].dropna().unique()
    colors = plt.cm.Set1(np.linspace(0, 1, len(families)))
    family_colors = dict(zip(families, colors))
    
    # Initial plots
    plot_fracture_rosette(df, ax1, family_colors)
    plot_fracture_stereonet(df, ax2, family_colors)
    
    # Create radio buttons for family selection
    ax_radio = plt.axes([0.05, 0.4, 0.15, 0.4])
    radio_labels = ['All'] + list(families)
    radio = RadioButtons(ax_radio, radio_labels)
    
    # Update function for radio buttons
    def update_fracture_plot(label):
        ax1.clear()
        ax2.clear()
        
        if label == 'All':
            filtered_df = df
        else:
            filtered_df = df[df['Familia'] == label]
        
        plot_fracture_rosette(filtered_df, ax1, family_colors)
        plot_fracture_stereonet(filtered_df, ax2, family_colors)
        
        fig.canvas.draw()
    
    radio.on_clicked(update_fracture_plot)
    
    return fig

def plot_fracture_rosette(df: pd.DataFrame, ax: plt.Axes, family_colors: Dict):
    """Plot fracture orientation rose diagram."""
    if 'Direccion_Buzamiento' not in df.columns:
        ax.text(0.5, 0.5, 'No direction data available', 
                transform=ax.transAxes, ha='center', va='center')
        return
    
    directions = df['Direccion_Buzamiento'].dropna()
    if directions.empty:
        ax.text(0.5, 0.5, 'No direction data available', 
                transform=ax.transAxes, ha='center', va='center')
        return
    
    # Convert to radians
    theta = np.radians(directions)
    
    # Create rose diagram
    bins = np.arange(0, 2*np.pi + np.pi/8, np.pi/8)  # 22.5-degree bins
    hist, _ = np.histogram(theta, bins=bins)
    
    # Plot histogram
    theta_centers = (bins[:-1] + bins[1:]) / 2
    ax = plt.subplot(121, projection='polar')
    ax.set_theta_zero_location('N')
    ax.set_theta_direction(-1)
    ax.bar(theta_centers, hist, width=np.pi/8, alpha=0.7)
    ax.set_title('Fracture Orientation Rose Diagram')

def plot_fracture_stereonet(df: pd.DataFrame, ax: plt.Axes, family_colors: Dict):
    """Plot fracture stereonet projection."""
    if not all(col in df.columns for col in ['Buzamiento', 'Direccion_Buzamiento']):
        ax.text(0.5, 0.5, 'Insufficient data for stereonet', 
                transform=ax.transAxes, ha='center', va='center')
        return
    
    # This would typically use mplstereonet, but we'll create a simplified version
    dips = df['Buzamiento'].dropna()
    dip_dirs = df['Direccion_Buzamiento'].dropna()
    
    if dips.empty or dip_dirs.empty:
        ax.text(0.5, 0.5, 'No fracture data available', 
                transform=ax.transAxes, ha='center', va='center')
        return
    
    # Simplified stereonet projection (normally would use proper stereonet projection)
    ax.set_aspect('equal')
    circle = plt.Circle((0, 0), 1, fill=False, color='black')
    ax.add_patch(circle)
    
    # Plot fracture poles (simplified projection)
    for i, (dip, dip_dir) in enumerate(zip(dips, dip_dirs)):
        # Simplified projection - in reality would use proper stereonet math
        x = (90 - dip) / 90 * np.cos(np.radians(dip_dir))
        y = (90 - dip) / 90 * np.sin(np.radians(dip_dir))
        
        family = df.iloc[i].get('Familia', 'Unknown') if i < len(df) else 'Unknown'
        color = family_colors.get(family, 'gray')
        ax.plot(x, y, 'o', color=color, markersize=6, alpha=0.7)
    
    ax.set_xlim(-1.1, 1.1)
    ax.set_ylim(-1.1, 1.1)
    ax.set_title('Fracture Stereonet (Simplified)')
    ax.axis('off')